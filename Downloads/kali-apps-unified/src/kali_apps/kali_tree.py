#!/usr/bin/env python3
"""
kali_tree.py — Phase 1 of the KALI platform

Takes a pairwise distance matrix CSV (output of pykali_hash.py or KALI_non-hash.py)
and produces:
  1. A Newick format tree file (.nwk)
  2. A self-contained interactive HTML tree viewer
  3. Optional: a static dendrogram PNG

Methods supported:
  --method upgma    Unweighted Pair Group Method with Arithmetic mean (default)
  --method nj       Neighbour Joining (more accurate for unequal rates)
  --method ward     Ward minimum variance (good for clustering)
  --method complete Complete linkage
  --method single   Single linkage
"""
from __future__ import annotations
import argparse
import json
import os
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from scipy.cluster import hierarchy
from scipy.spatial import distance


# ── Metadata helpers ──────────────────────────────────────────────────────────

def load_metadata(path: str, id_col: str = None,
                  name_col: str = None) -> pd.DataFrame:
    """
    Load a metadata CSV, TSV, or Excel (.xlsx) file.
    Supports KALI downloader Excel files automatically.
    """
    path = str(path)

    if path.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Metadata"] if "Metadata" in wb.sheetnames else wb.active
        rows    = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in rows[0]]
        df = pd.DataFrame(rows[1:], columns=headers).fillna("").astype(str)
    else:
        sep = "\t" if path.endswith((".tsv", ".txt")) else ","
        df  = pd.read_csv(path, sep=sep, dtype=str).fillna("")

    # Auto-detect KALI downloader format (CSV or Excel)
    kali_cols = {"Accession", "Organism"}
    if kali_cols.issubset(set(df.columns)):
        id_col   = id_col   or "Accession"
        name_col = name_col or "Organism"

        # Fill empty Genus from first word of Organism
        if "Genus" in df.columns:
            df["Genus"] = df.apply(
                lambda r: r["Genus"].strip() if r["Genus"].strip()
                          else r["Organism"].split()[0] if r["Organism"].strip()
                          else "Unknown",
                axis=1
            )
        else:
            df["Genus"] = df["Organism"].apply(
                lambda o: o.split()[0] if o.strip() else "Unknown"
            )

        # Short display: accession + first two words of organism
        def short_name(row):
            org   = str(row.get("Organism", ""))
            parts = org.split()
            short = " ".join(parts[:2]) if len(parts) >= 2 else org
            acc   = str(row.get("Accession",""))
            return f"{acc} | {short}" if short else acc
        df["_display_name"] = df.apply(short_name, axis=1)
        df["_group"] = df["Genus"]
    else:
        if id_col is None:   id_col   = df.columns[0]
        if name_col is None: name_col = df.columns[1] if len(df.columns) > 1 else id_col

    df = df.set_index(id_col)
    if "_display_name" not in df.columns:
        df["_display_name"] = df[name_col]
    return df


def extract_accession_from_label(label: str) -> str:
    """Extract NCBI accession from a KALI downloader filename stem.
    e.g. eco_shi_Escherichia_coli_NZ_CP110858 -> NZ_CP110858
    """
    import re
    m = re.search(r'([A-Z]{1,4}_?[A-Z]{0,2}\d{5,9})$', label)
    if m:
        return m.group(1)
    parts = label.split("_")
    if len(parts) >= 2:
        last_two = "_".join(parts[-2:])
        if re.match(r'^[A-Z]{1,4}\d{5,}$', last_two.split(".")[0]):
            return last_two
    return label


def build_rename_map(labels: list[str],
                     meta: pd.DataFrame) -> dict[str, str]:
    """
    Return {original_label: display_name} for every label that
    has a match in the metadata index.

    Matching strategy (in order):
      1. Exact match
      2. Accession extracted from KALI downloader filename stem
      3. Partial string match
      4. Keep original label
    """
    import re
    rename = {}
    for lbl in labels:
        # 1. Exact match
        if lbl in meta.index:
            rename[lbl] = meta.loc[lbl, "_display_name"]
            continue

        # 2. Extract accession from KALI filename stem
        #    e.g. eco_shi_Escherichia_coli_NZ_CP110858 -> NZ_CP110858
        m = re.search(r'([A-Z]{1,4}_?[A-Z]{0,2}\d{5,9})$', lbl)
        if m:
            acc = m.group(1)
            if acc in meta.index:
                rename[lbl] = meta.loc[acc, "_display_name"]
                continue
            # Also try without version suffix
            acc_base = acc.split(".")[0]
            if acc_base in meta.index:
                rename[lbl] = meta.loc[acc_base, "_display_name"]
                continue

        # 3. Partial match
        matches = [idx for idx in meta.index if idx in lbl or lbl in idx]
        if matches:
            rename[lbl] = meta.loc[matches[0], "_display_name"]
        else:
            rename[lbl] = lbl
    return rename


def apply_metadata_to_tree(tree_dict: dict,
                            meta: pd.DataFrame,
                            rename_map: dict[str, str]) -> dict:
    """
    Recursively walk the tree dict and:
      - Replace node names using rename_map
      - Attach metadata columns as extra fields for the HTML viewer
    Returns a new dict (does not mutate in place).
    """
    import copy
    node = copy.deepcopy(tree_dict)

    orig_name = node.get("name", "")
    if orig_name:
        node["original_id"]  = orig_name
        node["name"]         = rename_map.get(orig_name, orig_name)
        # Attach any extra metadata columns
        if orig_name in meta.index:
            row = meta.loc[orig_name]
            node["meta"] = {col: str(row[col])
                            for col in meta.columns
                            if not col.startswith("_")}

    if "children" in node:
        node["children"] = [
            apply_metadata_to_tree(child, meta, rename_map)
            for child in node["children"]
        ]
    return node


# ── Newick builder ────────────────────────────────────────────────────────────

def linkage_to_newick(linkage_matrix: np.ndarray, labels: list[str]) -> str:
    """Convert scipy linkage matrix to Newick string."""
    n = len(labels)
    # Each node is identified by an index; leaves are 0..n-1
    node_strings = {i: labels[i] for i in range(n)}

    for i, row in enumerate(linkage_matrix):
        left_idx  = int(row[0])
        right_idx = int(row[1])
        dist      = row[2]
        new_idx   = n + i

        left_str  = node_strings[left_idx]
        right_str = node_strings[right_idx]
        branch    = dist / 2.0   # half-distance as branch length

        node_strings[new_idx] = f"({left_str}:{branch:.6f},{right_str}:{branch:.6f})"

    root = node_strings[n + len(linkage_matrix) - 1]
    return root + ";"


# ── Neighbour joining ─────────────────────────────────────────────────────────

def neighbour_joining(dist_matrix: np.ndarray, labels: list[str]) -> str:
    """
    Simple neighbour-joining implementation.
    Returns a Newick string.
    """
    n = len(labels)
    d = dist_matrix.copy().astype(float)
    nodes = list(labels)
    edges = []   # (node_a, node_b, branch_a, branch_b)

    while len(nodes) > 2:
        size = len(nodes)
        # Q matrix
        row_sums = d.sum(axis=1)
        Q = np.zeros((size, size))
        for i in range(size):
            for j in range(size):
                if i != j:
                    Q[i, j] = (size - 2) * d[i, j] - row_sums[i] - row_sums[j]

        np.fill_diagonal(Q, np.inf)
        i, j = np.unravel_index(Q.argmin(), Q.shape)

        # Branch lengths
        branch_i = d[i, j] / 2 + (row_sums[i] - row_sums[j]) / (2 * (size - 2))
        branch_j = d[i, j] - branch_i

        # New node distances
        new_dists = [(d[i, k] + d[j, k] - d[i, j]) / 2
                     for k in range(size) if k != i and k != j]

        new_node = f"({nodes[i]}:{max(branch_i,0):.6f},{nodes[j]}:{max(branch_j,0):.6f})"

        # Remove i and j, add new node
        keep = [k for k in range(size) if k != i and k != j]
        new_d = np.zeros((len(keep) + 1, len(keep) + 1))
        for a, ka in enumerate(keep):
            for b, kb in enumerate(keep):
                new_d[a, b] = d[ka, kb]
            new_d[a, len(keep)] = new_dists[a]
            new_d[len(keep), a] = new_dists[a]

        nodes = [nodes[k] for k in keep] + [new_node]
        d = new_d

    if len(nodes) == 2:
        return f"({nodes[0]}:{d[0,1]/2:.6f},{nodes[1]}:{d[0,1]/2:.6f});"
    return nodes[0] + ";"


def bionj(dist_matrix: np.ndarray, labels: list[str]) -> str:
    """BIONJ — improved Neighbour Joining with variance-weighted distance updates.

    Produces more accurate branch lengths than plain NJ by using a
    weighted least-squares update that accounts for variance in distance
    estimates. Variance is estimated as proportional to distance (V_ij = d_ij).
    Returns a Newick string.
    """
    n     = len(labels)
    d     = dist_matrix.copy().astype(float)
    v     = d.copy()                        # variance matrix (V_ij = d_ij)
    nodes = list(labels)

    while len(nodes) > 2:
        size     = len(nodes)
        row_sums = d.sum(axis=1)

        # Q matrix — same as NJ
        Q = np.zeros((size, size))
        for i in range(size):
            for j in range(size):
                if i != j:
                    Q[i, j] = (size - 2) * d[i, j] - row_sums[i] - row_sums[j]
        np.fill_diagonal(Q, np.inf)
        i, j = np.unravel_index(Q.argmin(), Q.shape)

        # Branch lengths (same as NJ)
        branch_i = d[i, j] / 2 + (row_sums[i] - row_sums[j]) / (2 * (size - 2))
        branch_j = d[i, j] - branch_i

        # BIONJ weighted distance update
        new_dists = []
        new_vars  = []
        for k in range(size):
            if k == i or k == j:
                continue
            v_ij = v[i, j]
            if v_ij > 0:
                lam = 0.5 + (v[j, k] - v[i, k]) / (2 * v_ij)
                lam = max(0.0, min(1.0, lam))        # clamp to [0,1]
            else:
                lam = 0.5
            new_d = lam * (d[i, k] - branch_i) + (1 - lam) * (d[j, k] - branch_j)
            new_v = lam ** 2 * v[i, k] + (1 - lam) ** 2 * v[j, k] - lam * (1 - lam) * v_ij
            new_dists.append(max(new_d, 0.0))
            new_vars.append(max(new_v, 0.0))

        new_node = (
            f"({nodes[i]}:{max(branch_i,0):.6f},{nodes[j]}:{max(branch_j,0):.6f})"
        )

        keep  = [k for k in range(size) if k != i and k != j]
        sz2   = len(keep) + 1
        new_d = np.zeros((sz2, sz2))
        new_v = np.zeros((sz2, sz2))
        for a, ka in enumerate(keep):
            for b, kb in enumerate(keep):
                new_d[a, b] = d[ka, kb]
                new_v[a, b] = v[ka, kb]
            new_d[a, sz2 - 1] = new_dists[a]
            new_d[sz2 - 1, a] = new_dists[a]
            new_v[a, sz2 - 1] = new_vars[a]
            new_v[sz2 - 1, a] = new_vars[a]

        nodes = [nodes[k] for k in keep] + [new_node]
        d, v  = new_d, new_v

    if len(nodes) == 2:
        return f"({nodes[0]}:{d[0,1]/2:.6f},{nodes[1]}:{d[0,1]/2:.6f});"
    return nodes[0] + ";"


# ── Tree parser (Newick → dict for D3) ───────────────────────────────────────

def newick_to_dict(newick: str) -> dict:
    """Parse a Newick string into a nested dict for D3 hierarchy."""
    newick = newick.strip().rstrip(";")

    def parse(s: str) -> dict:
        s = s.strip()
        # Check for branch length
        branch = 0.0
        if ":" in s:
            # Find the last colon that is not inside parentheses
            depth = 0
            colon_pos = -1
            for idx, ch in enumerate(s):
                if ch == "(": depth += 1
                elif ch == ")": depth -= 1
                elif ch == ":" and depth == 0:
                    colon_pos = idx
            if colon_pos >= 0:
                try:
                    branch = float(s[colon_pos + 1:])
                except ValueError:
                    branch = 0.0
                s = s[:colon_pos]

        if s.startswith("(") and s.endswith(")"):
            inner = s[1:-1]
            # Split on commas at depth 0
            children_strs = []
            depth = 0
            start = 0
            for idx, ch in enumerate(inner):
                if ch == "(": depth += 1
                elif ch == ")": depth -= 1
                elif ch == "," and depth == 0:
                    children_strs.append(inner[start:idx])
                    start = idx + 1
            children_strs.append(inner[start:])
            children = [parse(c) for c in children_strs]
            return {"name": "", "branch_length": branch, "children": children}
        else:
            return {"name": s, "branch_length": branch}

    return parse(newick)


# ── HTML report ───────────────────────────────────────────────────────────────

def save_html_tree(tree_dict: dict, out_path: str, title: str,
                   method: str, n_taxa: int, metric: str,
                   meta=None, group_col=None,
                   dist_matrix=None, dist_labels=None,
                   linkage_matrix=None, all_labels=None) -> None:
    """Generate a self-contained React tree viewer.

    All data is injected as JSON into a <script id="kali-data"> tag so
    genome names, paths, and special characters never touch JS syntax.
    """
    # ── Build data payload ────────────────────────────────────────
    PALETTE = [
        "#1dc9a0","#f05050","#5b8df8","#f5a623","#9b72f8",
        "#00c8e0","#ff6b6b","#43e97b","#fa8231","#fd79a8",
        "#a29bfe","#fdcb6e","#6c5ce7","#55efc4","#e17055",
        "#74b9ff","#ff7675","#81ecec","#b2bec3","#636e72",
    ]

    group_colours: dict = {}
    if meta is not None and group_col and group_col in meta.columns:
        groups = sorted(meta[group_col].dropna().unique())
        group_colours = {g: PALETTE[i % len(PALETTE)] for i, g in enumerate(groups)}

    meta_lookup: dict = {}
    if meta is not None:
        for idx, row in meta.iterrows():
            meta_lookup[str(idx)] = {
                col: str(row[col])
                for col in meta.columns if not col.startswith("_")
            }

    payload = {
        "tree":         tree_dict,
        "title":        title,
        "method":       method,
        "metric":       metric,
        "nTaxa":        n_taxa,
        "groupColours": group_colours,
        "groupCol":     group_col or "",
        "metaLookup":   meta_lookup,
        "distMatrix":   dist_matrix  or [],
        "distLabels":   dist_labels  or [],
        "linkage":      linkage_matrix or [],
        "allLabels":    all_labels   or [],
    }
    data_json = json.dumps(payload)

    # ── Legend HTML (pure Python, no JS interaction needed) ───────
    if group_colours:
        items = "".join(
            '<span style="display:inline-flex;align-items:center;gap:5px;'
            'font-size:11px;margin-right:10px;">'
            '<span style="width:9px;height:9px;border-radius:50%;'
            'background:' + c + ';display:inline-block;"></span>'
            + str(g) + '</span>'
            for g, c in group_colours.items()
        )
        legend_block = (
            '<div style="padding:4px 1.5rem 8px;border-bottom:1px solid #2a3050;">'
            '<div style="font-size:9px;color:#6b7a9e;letter-spacing:.1em;'
            'text-transform:uppercase;margin-bottom:4px;">' + str(group_col) + '</div>'
            + items + '</div>'
        )
    else:
        legend_block = ""

    # ── HTML — data injected as JSON, React runs from CDN ─────────
    html = (
        "<!DOCTYPE html>\n"
        "<html lang=\"en\">\n"
        "<head>\n"
        "<meta charset=\"UTF-8\">\n"
        "<title>KALI Tree</title>\n"
        "<style>\n"
        "*{box-sizing:border-box;margin:0;padding:0;}"
        "body{font-family:\'IBM Plex Sans\',Arial,sans-serif;"
        "background:#0d0f14;color:#e2e8f8;}"
        "h1{font-size:1rem;font-weight:500;padding:.75rem 1.5rem .2rem;"
        "color:#1dc9a0;font-family:\'IBM Plex Mono\',monospace;}"
        ".meta-line{font-size:11px;color:#6b7a9e;padding:0 1.5rem .6rem;"
        "font-family:\'IBM Plex Mono\',monospace;}"
        ".cards{display:flex;gap:8px;padding:0 1.5rem .75rem;flex-wrap:wrap;}"
        ".card{background:#1c2133;border:1px solid #2a3050;border-radius:8px;"
        "padding:.4rem .8rem;}"
        ".cl{font-size:9px;color:#6b7a9e;text-transform:uppercase;"
        "letter-spacing:.08em;margin-bottom:1px;}"
        ".cv{font-size:.95rem;font-weight:500;}"
        ".toolbar{padding:.4rem 1.5rem .5rem;display:flex;gap:6px;"
        "flex-wrap:wrap;align-items:center;border-bottom:1px solid #2a3050;}"
        ".btn{padding:4px 11px;background:#1c2133;border:1px solid #2a3050;"
        "border-radius:5px;color:#6b7a9e;font-size:11px;cursor:pointer;"
        "font-family:monospace;transition:all .12s;}"
        ".btn:hover{border-color:#1dc9a0;color:#1dc9a0;}"
        ".btn.on{border-color:#5b8df8;color:#5b8df8;background:rgba(91,141,248,.1);}"
        "input[type=range]{accent-color:#1dc9a0;width:80px;vertical-align:middle;}"
        "#search{background:#1c2133;border:1px solid #2a3050;border-radius:5px;"
        "color:#e2e8f8;font-size:11px;font-family:monospace;padding:4px 9px;"
        "width:170px;outline:none;}"
        "#search:focus{border-color:#1dc9a0;}"
        ".sc{font-size:10px;color:#6b7a9e;}"
        "#tree-area{width:100%;overflow:auto;height:calc(100vh - 240px);}"
        "svg{display:block;cursor:grab;}"
        ".link{fill:none;stroke:#5b8df8;stroke-width:2.4;}"
        ".tip{position:fixed;background:#1c2133;border:1px solid #2a3050;"
        "border-radius:7px;padding:7px 11px;font-size:11px;font-family:monospace;"
        "color:#e2e8f8;pointer-events:none;opacity:0;transition:opacity .15s;"
        "z-index:100;max-width:260px;line-height:1.7;}"
        ".hi circle{fill:#f5a623!important;}"
        ".hi text{fill:#f5a623!important;font-weight:600;}"
        ".dim{opacity:.18;}"
        ".meta-upload{padding:4px 11px;background:#1c2133;border:1px solid #2a3050;"
        "border-radius:5px;color:#6b7a9e;font-size:11px;cursor:pointer;"
        "font-family:monospace;transition:all .12s;}"
        ".meta-upload:hover{border-color:#1dc9a0;color:#1dc9a0;}"
        "#meta-msg{font-size:10px;color:#1dc9a0;padding:3px 1.5rem;display:none;}"
        ".tab-bar{display:flex;gap:2px;padding:6px 1.5rem 0;border-bottom:1px solid #2a3050;}"
        ".tab-btn{padding:5px 16px;background:transparent;border:none;border-bottom:2px solid transparent;"
        "color:#6b7a9e;font-size:11px;font-family:monospace;cursor:pointer;transition:all .12s;}"
        ".tab-btn.active{color:#1dc9a0;border-bottom-color:#1dc9a0;}"
        ".tab-btn:hover{color:#e2e8f8;}"
        "#heatmap-area{width:100%;overflow:auto;height:calc(100vh - 240px);padding:1rem 1.5rem;}"
        "\n</style>\n"
        "</head>\n"
        "<body>\n"
        "<script id=\"kali-data\" type=\"application/json\">"
        + data_json +
        "</script>\n"
        "<div id=\"root\"></div>\n"
        "<div class=\"tip\" id=\"tip\"></div>\n"
        + legend_block +
        "<div id=\"meta-msg\"></div>\n"
        "<script src=\"https://cdnjs.cloudflare.com/ajax/libs/react/18.2.0/umd/react.production.min.js\"></script>\n"
        "<script src=\"https://cdnjs.cloudflare.com/ajax/libs/react-dom/18.2.0/umd/react-dom.production.min.js\"></script>\n"
        "<script src=\"https://cdnjs.cloudflare.com/ajax/libs/d3/7.8.5/d3.min.js\"></script>\n"
        "<script src=\"https://cdnjs.cloudflare.com/ajax/libs/babel-standalone/7.23.2/babel.min.js\"></script>\n"
        "<script type=\"text/babel\">\n"
        + r"""
const _data        = JSON.parse(document.getElementById('kali-data').textContent);
const TREE_DATA    = _data.tree;
const N_TAXA       = _data.nTaxa;
const LARGE        = N_TAXA > 80;
const INIT_SPACING = LARGE ? 12 : 22;
const INIT_FONT    = LARGE ? 8  : 11;
const INIT_LABELS  = !LARGE;

const PALETTE = [
  '#1dc9a0','#f05050','#5b8df8','#f5a623','#9b72f8',
  '#00c8e0','#ff6b6b','#43e97b','#fa8231','#fd79a8',
  '#a29bfe','#fdcb6e','#6c5ce7','#55efc4','#e17055',
  '#74b9ff','#ff7675','#81ecec','#b2bec3','#636e72',
];

function maxDepth(n){ return n.children ? 1+Math.max(...n.children.map(maxDepth)) : 0; }

// Collect max total branch length from root to leaf (for scale bar)
function maxRootDist(node, acc=0){
  if(!node.children) return acc + (node.branch_length||0);
  return Math.max(...node.children.map(c => maxRootDist(c, acc+(node.branch_length||0))));
}

function KaliTree() {
  const [layout,      setLayout]      = React.useState('rect');
  const [spacing,     setSpacing]     = React.useState(INIT_SPACING);
  const [fontSize,    setFontSize]    = React.useState(INIT_FONT);
  const [showLabels,  setShowLabels]  = React.useState(INIT_LABELS);
  const [search,      setSearch]      = React.useState('');
  const [proportional,setProportional]= React.useState(true);
  const [curvedLinks, setCurvedLinks] = React.useState(false);
  const [groupColours,setGroupColours]= React.useState(_data.groupColours);
  const [groupCol,    setGroupCol]    = React.useState(_data.groupCol);
  const [metaLookup,  setMetaLookup]  = React.useState(_data.metaLookup);
  const [treeData,    setTreeData]    = React.useState(TREE_DATA);
  const svgRef  = React.useRef(null);
  const zoomRef = React.useRef(null);

  React.useEffect(() => {
    if (!svgRef.current) return;
    const svg = d3.select(svgRef.current);
    const g   = svg.select('g.zoom-group');
    const z   = d3.zoom().scaleExtent([0.05, 20])
      .on('zoom', e => g.attr('transform', e.transform));
    svg.call(z);
    zoomRef.current = { svg, z };
  }, [layout, spacing, proportional, curvedLinks]);

  const resetZoom = () => {
    if (zoomRef.current)
      zoomRef.current.svg.transition().duration(350)
        .call(zoomRef.current.z.transform, d3.zoomIdentity);
  };

  const showTip = (e, d) => {
    const tip  = document.getElementById('tip');
    const name = d.data.name || '';
    const orig = d.data.original_id || '';
    const bl   = (d.data.branch_length || 0).toFixed(6);
    const meta = metaLookup[orig] || metaLookup[name] || {};
    let html = '<b>' + name + '</b>';
    if (orig && orig !== name) html += '<br><span style="color:#6b7a9e;font-size:10px;">' + orig + '</span>';
    html += '<br><span style="color:#6b7a9e;">branch length:</span> ' + bl;
    Object.entries(meta).forEach(([k,v]) => {
      html += '<br><span style="color:#6b7a9e;">' + k + ':</span> ' + v;
    });
    tip.innerHTML = html;
    tip.style.opacity = '1';
    tip.style.left = (e.clientX + 12) + 'px';
    tip.style.top  = (e.clientY - 8)  + 'px';
  };
  const hideTip = () => { document.getElementById('tip').style.opacity='0'; };

  // Colour by group if available, else by leaf/internal
  const nodeColour = (d) => {
    const orig = d.data.original_id || d.data.name;
    if (groupCol && metaLookup[orig] && metaLookup[orig][groupCol])
      return groupColours[metaLookup[orig][groupCol]] || '#5b8df8';
    return d.children ? '#2a3050' : '#5b8df8';
  };

  // For branches: propagate leaf colour up to parent if all children same group
  const branchColour = (link) => {
    const t = link.target;
    const orig = t.data.original_id || t.data.name;
    if (groupCol && metaLookup[orig] && metaLookup[orig][groupCol])
      return groupColours[metaLookup[orig][groupCol]] || '#5b8df8';
    return '#5b8df8';
  };

  const searchLC = search.toLowerCase().trim();
  const matchNode = (d) => {
    if (!searchLC) return false;
    const n = (d.data.name||'').toLowerCase();
    const o = (d.data.original_id||'').toLowerCase();
    const m = JSON.stringify(d.data.meta||{}).toLowerCase();
    return n.includes(searchLC)||o.includes(searchLC)||m.includes(searchLC);
  };

  const handleMetaFile = (e) => {
    const file = e.target.files[0]; if (!file) return;
    const reader = new FileReader();
    reader.onload = (ev) => applyMeta(ev.target.result, file.name);
    reader.readAsText(file); e.target.value='';
  };

  const applyMeta = (text, fname) => {
    const sep  = fname.endsWith('.tsv')||fname.endsWith('.txt') ? '\t' : ',';
    const rows = text.trim().split('\n').map(r=>r.split(sep).map(s=>s.trim().replace(/^"|"$/g,'')));
    if (rows.length < 2) return;
    const headers = rows[0];
    const nameCol = headers.length>1?1:0;
    const gi = headers.findIndex(h=>['group','genus','family','host','type','clade'].includes(h.toLowerCase()));
    const newLookup = {};
    rows.slice(1).forEach(row => {
      const id=row[0]; if(!id) return;
      const entry={_display:row[nameCol]||id};
      headers.forEach((h,i)=>{ if(i>0) entry[h]=row[i]||''; });
      newLookup[id]=entry;
    });
    const newGroupCol = gi>=0 ? headers[gi] : '';
    const newColours  = {};
    if (newGroupCol){
      const groups=[...new Set(rows.slice(1).map(r=>r[gi]).filter(Boolean))].sort();
      groups.forEach((g,i)=>{ newColours[g]=PALETTE[i%PALETTE.length]; });
    }
    const renamed = JSON.parse(JSON.stringify(treeData));
    const renameNode = (node) => {
      if (newLookup[node.original_id||node.name]){
        node.original_id=node.original_id||node.name;
        node.name=newLookup[node.original_id]._display||node.name;
      }
      if (node.children) node.children.forEach(renameNode);
    };
    renameNode(renamed);
    setTreeData(renamed); setGroupColours(newColours);
    setGroupCol(newGroupCol); setMetaLookup(newLookup);
    const msg=document.getElementById('meta-msg');
    msg.textContent='Loaded '+fname+' — '+Object.keys(newLookup).length+' records';
    msg.style.display='block';
    setTimeout(()=>{msg.style.display='none';},4000);
  };

  const buildTree = () => {
    const root   = d3.hierarchy(treeData);
    const leaves = root.leaves().length;
    const depth  = maxDepth(treeData);
    const totalDist = maxRootDist(treeData);

    if (layout === 'rect') {
      const mL=10, mR=240, mT=30, mB=30;
      const H = leaves * spacing + mT + mB;
      const innerW = Math.max(500, depth*110);
      const W = innerW + mL + mR;
      const scale = proportional && totalDist>0 ? innerW/totalDist : null;

      const clust = d3.cluster().size([H-mT-mB, innerW]);
      clust(root);

      // Assign y (horizontal) positions
      if (proportional && totalDist>0){
        // Walk top-down, accumulate true branch lengths
        root.each(d => {
          const bl = d.data.branch_length||0;
          d.y = d.parent ? d.parent.y + bl*scale : 0;
        });
        // Align all leaves to right edge
        const maxY = Math.max(...root.leaves().map(d=>d.y));
        root.leaves().forEach(d=>{ d.y=maxY; });
      } else {
        root.each(d => { d.y = d.depth * (innerW/Math.max(depth,1)); });
      }

      // Links — elbow or curved
      const links = root.links().map(link => {
        const sx=link.source.y, sy=link.source.x;
        const tx=link.target.y, ty=link.target.x;
        const pathD = curvedLinks
          ? `M${sx},${sy} C${(sx+tx)/2},${sy} ${(sx+tx)/2},${ty} ${tx},${ty}`
          : `M${tx},${ty}L${sx},${ty}L${sx},${sy}`;
        return { pathD, colour: branchColour(link),
                 key: link.target.data.name+link.target.x };
      });

      const nodes = root.descendants().map(d => {
        const isLeaf = !d.children&&!d._children;
        const matched= searchLC&&matchNode(d);
        const dimmed = searchLC&&!matchNode(d)&&isLeaf;
        return { d, isLeaf, matched, dimmed, tx:d.y, ty:d.x,
                 key:(d.data.name||'')+d.x+d.y };
      });

      // Scale bar — show at bottom, proportional mode only
      const scaleBar = (proportional && totalDist>0) ? {
        x1: mL, x2: mL + scale*totalDist*0.1,
        y: H - 10,
        label: (totalDist*0.1).toFixed(4)
      } : null;

      return { W, H, mL, mT, links, nodes, radial:false, scaleBar };

    } else {
      const SIZE = Math.max(500, leaves*spacing*1.1);
      const R    = SIZE/2;
      const W    = SIZE+300; const H=SIZE+100;
      const innerR = R-(LARGE?90:140);
      const scale  = proportional&&totalDist>0 ? innerR/totalDist : null;

      const clust = d3.cluster().size([2*Math.PI, innerR]);
      clust(root);

      if (proportional&&totalDist>0){
        root.each(d=>{
          const bl=d.data.branch_length||0;
          d.y = d.parent ? d.parent.y+bl*scale : 0;
        });
        const maxY=Math.max(...root.leaves().map(d=>d.y));
        root.leaves().forEach(d=>{d.y=maxY;});
      }

      const radLink = d3.linkRadial().angle(d=>d.x).radius(d=>d.y);
      const links = root.links().map(link => ({
        pathD: radLink(link),
        colour: branchColour(link),
        key: link.target.data.name+link.target.x
      }));

      const nodes = root.descendants().map(d=>{
        const isLeaf=!d.children&&!d._children;
        const matched=searchLC&&matchNode(d);
        const dimmed=searchLC&&!matchNode(d)&&isLeaf;
        const angle=d.x*180/Math.PI-90;
        const flip=d.x>=Math.PI;
        const textX=d.x<Math.PI?8:-8;
        const anchor=d.x<Math.PI?'start':'end';
        return {d,isLeaf,matched,dimmed,angle,flip,textX,anchor,
                tx:d.y, key:(d.data.name||'')+d.x+d.y};
      });

      return {W,H,cx:R+150,cy:R+50,links,nodes,radial:true,scaleBar:null};
    }
  };

  const tree = buildTree();
  const matchCount = searchLC ? tree.nodes.filter(n=>n.matched).length : 0;
  const [activeTab, setActiveTab] = React.useState('tree');

  return (
    <div>
      <h1>KALI phylogenetic tree</h1>
      <p className="meta-line">
        {_data.title} | method={_data.method} | metric={_data.metric} | taxa={N_TAXA}
      </p>
      <div className="cards">
        <div className="card"><div className="cl">Taxa</div><div className="cv">{N_TAXA}</div></div>
        <div className="card"><div className="cl">Method</div><div className="cv">{_data.method.toUpperCase()}</div></div>
        <div className="card"><div className="cl">Metric</div><div className="cv">{_data.metric}</div></div>
        <div className="card"><div className="cl">Groups</div>
          <div className="cv">{Object.keys(_data.groupColours).length||'—'}</div></div>
      </div>

      {/* Tab bar */}
      <div className="tab-bar">
        <button className={'tab-btn'+(activeTab==='tree'?' active':'')}
                onClick={()=>setActiveTab('tree')}>🌿 Tree</button>
        <button className={'tab-btn'+(activeTab==='heatmap'?' active':'')}
                onClick={()=>setActiveTab('heatmap')}>🗂 Heatmap</button>
      </div>

      {/* Tree tab */}
      {activeTab==='tree' && <>
      <div className="toolbar">
        <button className={'btn'+(layout==='rect'?' on':'')} onClick={()=>setLayout('rect')}>Rectangular</button>
        <button className={'btn'+(layout==='rad'?' on':'')}  onClick={()=>setLayout('rad')}>Radial</button>
        <button className={'btn'+(proportional?' on':'')} onClick={()=>setProportional(v=>!v)}>
          {proportional?'Proportional':'Cladogram'}
        </button>
        <button className={'btn'+(curvedLinks?' on':'')} onClick={()=>setCurvedLinks(v=>!v)}>
          {curvedLinks?'Curved':'Elbow'}
        </button>
        <label style={{fontSize:11,color:'#6b7a9e'}}>Spacing </label>
        <input type="range" min="6" max="40" value={spacing} onChange={e=>setSpacing(+e.target.value)}/>
        <label style={{fontSize:11,color:'#6b7a9e'}}>Font </label>
        <input type="range" min="6" max="16" value={fontSize} onChange={e=>setFontSize(+e.target.value)}/>
        <input id="search" placeholder="Search nodes…"
               value={search} onChange={e=>setSearch(e.target.value)}/>
        {searchLC && <span className="sc">{matchCount} match{matchCount!==1?'es':''}</span>}
        <button className="btn" onClick={()=>setShowLabels(v=>!v)}>
          {showLabels?'Hide labels':'Show labels'}
        </button>
        <button className="btn" onClick={resetZoom}>Reset zoom</button>
        <label className="meta-upload" htmlFor="meta-file">Upload metadata</label>
        <input id="meta-file" type="file" accept=".csv,.tsv,.txt"
               style={{display:'none'}} onChange={handleMetaFile}/>
        <button className="btn" onClick={()=>{
          const s=document.querySelector('#tree-area svg');
          if(!s) return;
          let svgStr = s.outerHTML
            .replace(/&nbsp;/g, '&#160;')
            .replace(/&mdash;/g, '&#8212;')
            .replace(/&ndash;/g, '&#8211;')
            .replace(/&copy;/g, '&#169;')
            .replace(/&amp;nbsp;/g, '&#160;');
          // Add xmlns if missing so Preview and Illustrator can open it
          if (!svgStr.includes('xmlns='))
            svgStr = svgStr.replace('<svg ', '<svg xmlns="http://www.w3.org/2000/svg" ');
          const blob = new Blob(
            ['<?xml version="1.0" encoding="UTF-8"?>\n' + svgStr],
            {type:'image/svg+xml;charset=utf-8'}
          );
          const a=document.createElement('a');
          a.href=URL.createObjectURL(blob);
          a.download='kali_tree.svg'; a.click();
        }}>SVG ↓</button>
      </div>
      <div id="tree-area">
        {!tree.radial ? (
          <svg ref={svgRef} width={tree.W} height={tree.H}>
            <g className="zoom-group" transform={'translate('+tree.mL+','+tree.mT+')'}>
              {tree.links.map(l => (
                <path key={l.key} d={l.pathD}
                      style={{fill:'none', stroke: l.colour,
                              strokeWidth: l.colour==='#5b8df8'?2.4:1.8,
                              opacity: l.colour==='#5b8df8'?1:0.85}}/>
              ))}
              {tree.nodes.map(({d,isLeaf,matched,dimmed,tx,ty,key}) => (
                <g key={key}
                   className={'node'+(matched?' hi':'')+(dimmed?' dim':'')}
                   transform={'translate('+tx+','+ty+')'}
                   onMouseMove={e=>showTip(e,d)}
                   onMouseOut={hideTip}
                   onClick={!isLeaf ? ()=>{
                     if(d._children){d.children=d._children;d._children=null;}
                     else{d._children=d.children;d.children=null;}
                     setSpacing(v=>v);
                   } : undefined}>
                  <circle r={isLeaf?(LARGE?4:7):4}
                          fill={nodeColour(d)}
                          stroke={isLeaf?nodeColour(d):'#1dc9a0'}
                          strokeWidth="1.2"/>
                  {isLeaf && (
                    <text dx={8} dy="0.35em"
                          style={{fontSize:fontSize+'px',
                                  display:showLabels?'block':'none',
                                  fontFamily:'IBM Plex Mono,monospace',
                                  fill: nodeColour(d)==='#5b8df8' ? '#e2e8f8' : nodeColour(d)}}>
                      {d.data.name}
                    </text>
                  )}
                </g>
              ))}
              {/* Scale bar */}
              {tree.scaleBar && (
                <g transform={'translate(0,'+(tree.H-tree.mT-tree.mB-5)+')'}>
                  <line x1={tree.scaleBar.x1} x2={tree.scaleBar.x2}
                        y1={0} y2={0} stroke="#6b7a9e" strokeWidth="1.5"/>
                  <line x1={tree.scaleBar.x1} x2={tree.scaleBar.x1}
                        y1={-4} y2={4} stroke="#6b7a9e" strokeWidth="1.5"/>
                  <line x1={tree.scaleBar.x2} x2={tree.scaleBar.x2}
                        y1={-4} y2={4} stroke="#6b7a9e" strokeWidth="1.5"/>
                  <text x={(tree.scaleBar.x1+tree.scaleBar.x2)/2} y={14}
                        textAnchor="middle"
                        style={{fontSize:'10px',fill:'#6b7a9e',fontFamily:'monospace'}}>
                    {tree.scaleBar.label}
                  </text>
                </g>
              )}
            </g>
          </svg>
        ) : (
          <svg ref={svgRef} width={tree.W} height={tree.H}>
            <g className="zoom-group" transform={'translate('+tree.cx+','+tree.cy+')'}>
              {tree.links.map(l => (
                <path key={l.key} d={l.pathD}
                      style={{fill:'none', stroke: l.colour,
                              strokeWidth: l.colour==='#5b8df8'?2.4:1.8,
                              opacity: l.colour==='#5b8df8'?1:0.85}}/>
              ))}
              {tree.nodes.map(({d,isLeaf,matched,dimmed,angle,flip,textX,anchor,tx,key}) => (
                <g key={key}
                   className={'node'+(matched?' hi':'')+(dimmed?' dim':'')}
                   transform={'rotate('+angle+') translate('+tx+',0)'}
                   onMouseMove={e=>showTip(e,d)}
                   onMouseOut={hideTip}>
                  <circle r={isLeaf?(LARGE?4:7):3}
                          fill={nodeColour(d)}
                          stroke={isLeaf?nodeColour(d):'#1dc9a0'}
                          strokeWidth="1.2"/>
                  {isLeaf && (
                    <text x={textX} dy="0.31em"
                          textAnchor={anchor}
                          transform={flip?'rotate(180)':undefined}
                          style={{fontSize:fontSize+'px',
                                  display:showLabels?'block':'none',
                                  fontFamily:'IBM Plex Mono,monospace',
                                  fill: nodeColour(d)==='#5b8df8' ? '#e2e8f8' : nodeColour(d)}}>
                      {d.data.name}
                    </text>
                  )}
                </g>
              ))}
            </g>
          </svg>
        )}
      </div>
    </>}

    {/* Heatmap tab */}
    {activeTab==='heatmap' && <KaliHeatmap/>}
    </div>
  );
}

// ── Heatmap + dendrogram component ───────────────────────────────────────────
function KaliHeatmap() {
  const mat    = _data.distMatrix || [];
  const labels = _data.distLabels || [];
  const lm     = _data.linkage    || [];
  const n      = labels.length;
  const gc     = _data.groupColours || {};
  const gcol   = _data.groupCol     || '';
  const ml     = _data.metaLookup   || {};

  const [hover, setHover] = React.useState(null);

  if (!mat.length || n === 0)
    return <div style={{padding:'2rem',color:'#6b7a9e'}}>No distance matrix data.</div>;

  const CELL    = Math.max(5, Math.min(20, Math.floor(520/n)));
  const DEND_W  = 110;
  const LABEL_W = 160;
  const PAD     = 8;
  const maxV    = Math.max(...mat.flat().filter(v=>v>0));

  const cellColor = (v) => {
    const t = Math.min(v / (maxV + 1e-9), 1);
    const r = Math.round(t*230 + (1-t)*29);
    const g = Math.round(t*60  + (1-t)*201);
    const b = Math.round(t*60  + (1-t)*160);
    return `rgb(${r},${g},${b})`;
  };

  const groupColor = (lb) => {
    const e = ml[lb] || {};
    return (gcol && e[gcol] && gc[e[gcol]]) ? gc[e[gcol]] : '#5b8df8';
  };

  // Build dendrogram lines from linkage matrix
  const nodeY = {};
  labels.forEach((_,i) => { nodeY[i] = PAD + (i + 0.5) * CELL; });
  const nodeX = {};
  labels.forEach((_,i) => { nodeX[i] = 0; });
  const dendLines = [];

  lm.forEach(([li, lj, dist], stepIdx) => {
    const newIdx = n + stepIdx;
    const y1   = nodeY[Math.round(li)] ?? PAD + (li+0.5)*CELL;
    const y2   = nodeY[Math.round(lj)] ?? PAD + (lj+0.5)*CELL;
    const yMid = (y1+y2)/2;
    nodeY[newIdx] = yMid;
    const x    = (dist/maxV) * (DEND_W-12);
    nodeX[newIdx] = x;
    const cx1  = nodeX[Math.round(li)] ?? 0;
    const cx2  = nodeX[Math.round(lj)] ?? 0;
    // horizontal line for child i
    dendLines.push({x1: DEND_W-cx1, y1, x2: DEND_W-x, y2: y1});
    // horizontal line for child j
    dendLines.push({x1: DEND_W-cx2, y1: y2, x2: DEND_W-x, y2});
    // vertical connector
    dendLines.push({x1: DEND_W-x, y1, x2: DEND_W-x, y2});
  });

  const totalH = PAD*2 + n*CELL;
  const totalW = DEND_W + n*CELL + LABEL_W + 10;

  return (
    <div id="heatmap-area">
      <div style={{display:'flex',alignItems:'center',gap:14,marginBottom:8,flexWrap:'wrap'}}>
        <span style={{fontSize:11,color:'#6b7a9e',fontFamily:'monospace'}}>
          {n}×{n} distance matrix · ordered by {_data.method.toUpperCase()} linkage
        </span>
        <div style={{display:'flex',alignItems:'center',gap:4,fontSize:10,color:'#6b7a9e'}}>
          <span>Low</span>
          <svg width={80} height={8}>
            {Array.from({length:80},(_,i)=>(
              <rect key={i} x={i} y={0} width={1} height={8}
                    fill={cellColor(i/80*maxV)}/>
            ))}
          </svg>
          <span>High</span>
        </div>
        {hover && (
          <span style={{fontSize:10,color:'#e2e8f8',fontFamily:'monospace',
                        background:'#1c2133',padding:'2px 8px',borderRadius:4,
                        border:'1px solid #2a3050'}}>
            {hover.r} ↔ {hover.c} &nbsp;|&nbsp; {hover.v.toFixed(5)}
          </span>
        )}
      </div>

      <svg width={totalW} height={totalH + CELL*3}>

        {/* Dendrogram */}
        <g>
          {dendLines.map((l,i)=>(
            <line key={i} x1={l.x1} y1={l.y1} x2={l.x2} y2={l.y2}
                  stroke='#1dc9a0' strokeWidth={1} opacity={0.75}/>
          ))}
        </g>

        {/* Row labels */}
        <g transform={`translate(${DEND_W+2},0)`}>
          {labels.map((lb,i)=>(
            <g key={i}>
              <circle cx={-6} cy={PAD+(i+0.5)*CELL} r={3} fill={groupColor(lb)}/>
              <text x={-12} y={PAD+(i+0.5)*CELL} dy="0.35em"
                    textAnchor="end"
                    style={{fontSize:Math.min(CELL-1,10)+'px',
                            fill:'#e2e8f8',fontFamily:'IBM Plex Mono,monospace'}}>
                {lb.length>20 ? lb.slice(0,18)+'…' : lb}
              </text>
            </g>
          ))}
        </g>

        {/* Heatmap cells */}
        <g transform={`translate(${DEND_W+2},${PAD})`}>
          {mat.map((row,ri)=>
            row.map((val,ci)=>(
              <rect key={`${ri}-${ci}`}
                    x={ci*CELL} y={ri*CELL}
                    width={CELL-1} height={CELL-1}
                    fill={ri===ci?'#161a24':cellColor(val)}
                    stroke={ri===ci?'#1dc9a0':'none'}
                    strokeWidth={0.5}
                    onMouseEnter={()=>setHover({r:labels[ri],c:labels[ci],v:val})}
                    onMouseLeave={()=>setHover(null)}/>
            ))
          )}
        </g>

        {/* Column labels (rotated) */}
        <g transform={`translate(${DEND_W+2},${PAD+n*CELL+4})`}>
          {labels.map((lb,i)=>(
            <text key={i}
                  transform={`translate(${i*CELL+CELL/2},0) rotate(-50)`}
                  textAnchor="end"
                  style={{fontSize:Math.min(CELL-1,9)+'px',
                          fill:'#6b7a9e',fontFamily:'IBM Plex Mono,monospace'}}>
              {lb.length>14 ? lb.slice(0,12)+'…' : lb}
            </text>
          ))}
        </g>
      </svg>
    </div>
  );
}

ReactDOM.createRoot(document.getElementById('root')).render(<KaliTree/>);
"""
    +
        "\n</script>\n"
        "</body>\n"
        "</html>\n"
    )

    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html)


def save_dendrogram_png(linkage_matrix: np.ndarray, labels: list[str],
                         out_path: str, method: str) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(10, max(4, len(labels) * 0.4)))
        hierarchy.dendrogram(
            linkage_matrix, labels=labels, orientation="left",
            leaf_font_size=9, ax=ax
        )
        ax.set_title(f"KALI dendrogram — {method.upper()}", fontsize=11)
        ax.set_xlabel("Distance")
        plt.tight_layout()
        plt.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close()
    except ImportError:
        print("  matplotlib not available — skipping PNG")


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Build a phylogenetic tree from a KALI distance matrix CSV"
    )
    p.add_argument("-i", "--input",  required=True,
                   help="Distance matrix CSV (from pykali_hash.py or KALI_non-hash.py)")
    p.add_argument("-o", "--output", default=None,
                   help="Output base name (default: input filename without .csv)")
    p.add_argument("-m", "--method",
                   choices=["upgma", "nj", "bionj", "ward", "complete", "single", "average"],
                   default="upgma",
                   help="Tree building method (default: upgma)")
    p.add_argument("--metric", default="cosine",
                   help="Distance metric label for display (default: cosine)")
    p.add_argument("--png", action="store_true",
                   help="Also save a static dendrogram PNG")
    p.add_argument("--metadata", default=None,
                   help="CSV/TSV metadata file for renaming nodes and colouring")
    p.add_argument("--id-col", default=None,
                   help="Column in metadata matching distance matrix labels (default: first col)")
    p.add_argument("--name-col", default=None,
                   help="Column to use as display name on the tree (default: second col)")
    p.add_argument("--group-col", default=None,
                   help="Column to use for node colour grouping (e.g. 'group', 'host')")
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args()


def main() -> None:
    args  = parse_args()
    infile = Path(args.input)
    out_base = args.output or str(infile.with_suffix(""))

    # Load distance matrix
    df = pd.read_csv(args.input, index_col=0)
    labels = list(df.index)
    n = len(labels)

    if args.verbose:
        print(f"Loaded: {infile.name}  |  {n} taxa")
        print(f"Labels: {', '.join(labels[:5])}{'…' if n>5 else ''}")

    # Load metadata if provided
    meta       = None
    rename_map = {lbl: lbl for lbl in labels}   # identity by default
    if args.metadata:
        meta = load_metadata(args.metadata,
                             id_col=args.id_col,
                             name_col=args.name_col)
        rename_map = build_rename_map(labels, meta)
        matched = sum(1 for v in rename_map.values()
                      if v != rename_map.get(v, v) or True)
        n_renamed = sum(1 for k, v in rename_map.items() if k != v)
        if args.verbose:
            print(f"Metadata: {Path(args.metadata).name}  |  "
                  f"{n_renamed}/{n} labels renamed")
            for orig, disp in list(rename_map.items())[:5]:
                if orig != disp:
                    print(f"  {orig!r}  →  {disp!r}")

    mat = df.values.astype(float)
    # Ensure symmetry and zero diagonal
    mat = (mat + mat.T) / 2
    np.fill_diagonal(mat, 0)

    # Build tree
    if args.verbose:
        print(f"Building tree using: {args.method.upper()}")

    # Use display names in Newick if metadata provided
    display_labels = [rename_map.get(lbl, lbl) for lbl in labels]

    if args.method == "nj":
        newick = neighbour_joining(mat, display_labels)
        condensed = distance.squareform(mat)
        lm = hierarchy.linkage(condensed, method="average")
    elif args.method == "bionj":
        newick = bionj(mat, display_labels)
        condensed = distance.squareform(mat)
        lm = hierarchy.linkage(condensed, method="average")
    else:
        scipy_method = "average" if args.method == "upgma" else args.method
        condensed = distance.squareform(mat)
        lm = hierarchy.linkage(condensed, method=scipy_method)
        newick = linkage_to_newick(lm, display_labels)

    # Save Newick
    nwk_path = out_base + ".nwk"
    Path(nwk_path).write_text(newick)
    print(f"Saved Newick:      {nwk_path}")

    # Build tree dict and attach metadata
    tree_dict = newick_to_dict(newick)
    if meta is not None:
        tree_dict = apply_metadata_to_tree(tree_dict, meta, rename_map)

    html_path = out_base + "_tree.html"

    # Compute leaf order from linkage for heatmap
    leaf_order = hierarchy.leaves_list(lm).tolist()
    ordered_labels = [display_labels[i] for i in leaf_order]
    ordered_mat    = mat[np.ix_(leaf_order, leaf_order)]
    # Build linkage steps for D3 dendrogram in heatmap
    lm_list = lm.tolist()

    save_html_tree(
        tree_dict, html_path,
        title=infile.stem,
        method=args.method,
        n_taxa=n,
        metric=args.metric,
        meta=meta,
        group_col=args.group_col,
        dist_matrix=ordered_mat.tolist(),
        dist_labels=ordered_labels,
        linkage_matrix=lm_list,
        all_labels=display_labels,
    )
    print(f"Saved HTML tree:   {html_path}")

    # Optional PNG
    if args.png:
        png_path = out_base + "_dendrogram.png"
        save_dendrogram_png(lm, display_labels, png_path, args.method)
        print(f"Saved dendrogram:  {png_path}")

    if args.verbose:
        print(f"\nNewick preview:\n  {newick[:120]}{'…' if len(newick)>120 else ''}")


if __name__ == "__main__":
    main()
