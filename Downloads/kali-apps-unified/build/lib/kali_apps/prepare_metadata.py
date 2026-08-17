#!/usr/bin/env python3
"""
prepare_metadata.py — Fix KALI downloader metadata for classifier training

NCBI does not always fill in the Genus column. This script:
  1. Auto-fills Genus from the Organism name (first word)
  2. Infers Family from genus using a built-in lookup table
  3. Removes duplicate accessions
  4. Removes sequences shorter than a minimum length
  5. Saves a clean CSV ready for kali_classifier.py

USAGE
-----
python prepare_metadata.py \
  --input  eco_shi_merged_metadata.xlsx \
  --output enterobac_labels.csv \
  --label-col Genus \
  --min-length 100000

# For multi-file merge (one Excel per genus download):
python prepare_metadata.py \
  --input  ecoli_metadata.xlsx shigella_metadata.xlsx salmonella_metadata.xlsx \
  --output enterobac_labels.csv \
  --label-col Genus
"""
from __future__ import annotations
import argparse
import re
import sys
from pathlib import Path

import pandas as pd

# ── Built-in genus → family lookup ───────────────────────────────────────────
GENUS_TO_FAMILY = {
    # Enterobacteriaceae
    "Escherichia":    "Enterobacteriaceae",
    "Shigella":       "Enterobacteriaceae",
    "Salmonella":     "Enterobacteriaceae",
    "Klebsiella":     "Enterobacteriaceae",
    "Enterobacter":   "Enterobacteriaceae",
    "Citrobacter":    "Enterobacteriaceae",
    "Serratia":       "Enterobacteriaceae",
    "Proteus":        "Enterobacteriaceae",
    "Yersinia":       "Enterobacteriaceae",
    "Cronobacter":    "Enterobacteriaceae",
    "Hafnia":         "Enterobacteriaceae",
    "Pantoea":        "Enterobacteriaceae",
    # Pseudomonadaceae
    "Pseudomonas":    "Pseudomonadaceae",
    # Staphylococcaceae
    "Staphylococcus": "Staphylococcaceae",
    # Streptococcaceae
    "Streptococcus":  "Streptococcaceae",
    "Lactococcus":    "Streptococcaceae",
    # Mycobacteriaceae
    "Mycobacterium":  "Mycobacteriaceae",
    "Mycobacteroides":"Mycobacteriaceae",
    # Bacillaceae
    "Bacillus":       "Bacillaceae",
    "Geobacillus":    "Bacillaceae",
    # Clostridiaceae
    "Clostridium":    "Clostridiaceae",
    # Neisseriaceae
    "Neisseria":      "Neisseriaceae",
    # Flaviviridae genera (viral)
    "Flavivirus":     "Flaviviridae",
    "Orthoflavivirus":"Flaviviridae",
    "Hepacivirus":    "Flaviviridae",
    "Pegivirus":      "Flaviviridae",
    "Pestivirus":     "Flaviviridae",
    "Orthopestivirus":"Flaviviridae",
}


def load_file(path: str) -> pd.DataFrame:
    """Load CSV or Excel metadata file."""
    path = str(path)
    if path.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Metadata"] if "Metadata" in wb.sheetnames else wb.active
            rows    = list(ws.iter_rows(values_only=True))
            headers = [str(h) if h else "" for h in rows[0]]
            df = pd.DataFrame(rows[1:], columns=headers)
        except Exception as e:
            raise ValueError(f"Cannot read Excel {path}: {e}")
    else:
        df = pd.read_csv(path, encoding="utf-8-sig")
    return df.fillna("").astype(str)


def fix_metadata(df: pd.DataFrame,
                 min_length: int = 0,
                 verbose: bool = True) -> pd.DataFrame:
    """
    Fix NCBI metadata:
      - Fill Genus from Organism first word
      - Fill Family from genus lookup
      - Remove duplicates by Accession
      - Remove sequences below min_length
    """
    original_n = len(df)

    # Ensure required columns exist
    for col in ["Accession", "Organism", "Genus", "Family",
                "Length (bp)", "Filename"]:
        if col not in df.columns:
            df[col] = ""

    # Fill Genus from Organism if empty
    mask_no_genus = df["Genus"].str.strip() == ""
    df.loc[mask_no_genus, "Genus"] = df.loc[mask_no_genus, "Organism"].apply(
        lambda o: o.strip().split()[0] if o.strip() else "Unknown"
    )
    n_filled_genus = mask_no_genus.sum()

    # Fill Family from lookup if empty
    mask_no_family = df["Family"].str.strip() == ""
    df.loc[mask_no_family, "Family"] = df.loc[mask_no_family, "Genus"].apply(
        lambda g: GENUS_TO_FAMILY.get(g.strip(), "")
    )

    # Remove duplicates
    before_dedup = len(df)
    df = df.drop_duplicates(subset=["Accession"])
    n_dupes = before_dedup - len(df)

    # Filter by length
    n_short = 0
    if min_length > 0:
        before_len = len(df)
        try:
            df["_len"] = pd.to_numeric(df["Length (bp)"], errors="coerce").fillna(0)
            df = df[df["_len"] >= min_length].drop(columns=["_len"])
        except Exception:
            pass
        n_short = before_len - len(df)

    if verbose:
        print(f"  Original rows:       {original_n}")
        print(f"  Genus filled:        {n_filled_genus}")
        print(f"  Duplicates removed:  {n_dupes}")
        print(f"  Too short removed:   {n_short}")
        print(f"  Final rows:          {len(df)}")
        print()
        print("  Genus distribution:")
        for g, c in df["Genus"].value_counts().items():
            print(f"    {g:<20} {c}")

    return df.reset_index(drop=True)


def main():
    p = argparse.ArgumentParser(
        description="Fix NCBI metadata and prepare for KALI classifier training"
    )
    p.add_argument("--input",   nargs="+", required=True,
                   help="Input metadata file(s) — Excel or CSV")
    p.add_argument("--output",  required=True,
                   help="Output CSV path")
    p.add_argument("--label-col", default="Genus",
                   help="Column to use as class label (default: Genus)")
    p.add_argument("--min-length", type=int, default=0,
                   help="Minimum sequence length in bp (default: 0 = no filter)")
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()

    print(f"\n── Prepare Metadata for KALI Classifier ──────────────")

    # Load and merge all input files
    dfs = []
    for path in args.input:
        if not Path(path).exists():
            print(f"  WARNING: file not found: {path}")
            continue
        print(f"  Loading: {path}")
        dfs.append(load_file(path))

    if not dfs:
        print("No valid input files found.")
        sys.exit(1)

    df = pd.concat(dfs, ignore_index=True)
    print(f"  Loaded {len(df)} total rows from {len(dfs)} file(s)\n")

    # Fix
    df = fix_metadata(df, min_length=args.min_length, verbose=True)

    # Save
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(args.output, index=False)
    print(f"\n  Saved: {args.output}")
    print(f"\nTo train the classifier:")
    print(f"  python kali_classifier.py train \\")
    print(f"    --genomes /path/to/genome/folder \\")
    print(f"    --metadata {args.output} \\")
    print(f"    --accession-col Accession \\")
    print(f"    --label-col {args.label_col} \\")
    print(f"    -k 3 4 5 \\")
    print(f"    -o models/enterobac_model.pkl")
    print()


if __name__ == "__main__":
    main()
