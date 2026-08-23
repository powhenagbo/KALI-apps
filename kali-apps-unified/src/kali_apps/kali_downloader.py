#!/usr/bin/env python3
"""
kali_downloader.py — NCBI Genome Downloader for KALI

Downloads genome sequences from NCBI by:
  - Taxonomy ID (txid)
  - Organism name / family / genus
  - Accession list

Also generates an Excel metadata file with accession, organism,
family, genus, length, GC content, and collection info.

USAGE
-----
# Download all complete Flaviviridae genomes (up to 200)
python kali_downloader.py \
  --taxon "Flaviviridae" \
  --limit 200 \
  --complete \
  -o /path/to/output_folder

# Download by taxid (11050 = Flaviviridae)
python kali_downloader.py \
  --taxid 11050 \
  --limit 100 \
  -o /path/to/output_folder

# Download specific accessions
python kali_downloader.py \
  --accessions NC_001474 NC_001477 NC_002640 \
  -o /path/to/output_folder
"""
from __future__ import annotations

import argparse
import re
import sys
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── NCBI helpers ──────────────────────────────────────────────────────────────

BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
HEADERS  = {"User-Agent": "KALI-Downloader/1.0 (mailto:research@ualr.edu)"}


def ncbi_get(url: str, retries: int = 3) -> bytes:
    """Fetch a URL with retries and rate limiting."""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=30) as r:
                return r.read()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                raise
    return b""


def esearch(term: str, db: str = "nucleotide",
            limit: int = 100, usehistory: bool = True) -> dict:
    """Search NCBI and return IDs + WebEnv for large fetches."""
    url = (f"{BASE_URL}/esearch.fcgi"
           f"?db={db}&term={urllib.parse.quote(term)}"
           f"&retmax={limit}&retmode=json"
           f"{'&usehistory=y' if usehistory else ''}")
    import json
    data = json.loads(ncbi_get(url))
    return data.get("esearchresult", {})


def efetch_fasta(ids: list[str], db: str = "nucleotide") -> str:
    """Fetch FASTA sequences for a list of IDs."""
    id_str = ",".join(ids)
    url    = (f"{BASE_URL}/efetch.fcgi"
              f"?db={db}&id={id_str}&rettype=fasta&retmode=text")
    return ncbi_get(url).decode("utf-8", errors="replace")


def efetch_summary(ids: list[str], db: str = "nucleotide") -> list[dict]:
    """Fetch DocSummary for metadata extraction."""
    import json
    id_str = ",".join(ids)
    url    = (f"{BASE_URL}/esummary.fcgi"
              f"?db={db}&id={id_str}&retmode=json")
    data   = json.loads(ncbi_get(url))
    result = data.get("result", {})
    return [result[uid] for uid in result.get("uids", []) if uid in result]


def taxon_to_txid(name: str) -> str | None:
    """Convert organism/family name to NCBI taxonomy ID."""
    result = esearch(f"{name}[All Names]", db="taxonomy", limit=1, usehistory=False)
    ids = result.get("idlist", [])
    return ids[0] if ids else None


def gc_content(seq: str) -> float:
    seq = seq.upper().replace("\n","").replace(" ","")
    gc  = seq.count("G") + seq.count("C")
    tot = sum(seq.count(b) for b in "ACGT")
    return round(gc / tot * 100, 2) if tot > 0 else 0.0


# ── Excel metadata writer ─────────────────────────────────────────────────────

def write_excel_metadata(records: list[dict], out_path: str) -> None:
    """Write a formatted Excel metadata file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metadata"

    # Header style
    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alt row fill
    alt_fill  = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
    body_font = Font(name="Arial", size=10)
    body_align= Alignment(vertical="center")

    headers = [
        "Accession", "Organism", "Family", "Genus",
        "Length (bp)", "GC %", "Title / Description",
        "Collection Date", "Country", "Host", "Filename"
    ]
    col_widths = [14, 28, 18, 18, 12, 8, 48, 14, 16, 18, 30]

    # Write headers
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Write data rows
    for ri, rec in enumerate(records, 2):
        fill = alt_fill if ri % 2 == 0 else None
        row_vals = [
            rec.get("accession", ""),
            rec.get("organism", ""),
            rec.get("family", ""),
            rec.get("genus", ""),
            rec.get("length", ""),
            rec.get("gc", ""),
            rec.get("title", ""),
            rec.get("collection_date", ""),
            rec.get("country", ""),
            rec.get("host", ""),
            rec.get("filename", ""),
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = body_font
            cell.alignment = body_align
            cell.border    = cell_border
            if fill:
                cell.fill = fill

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "KALI Download Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

    ws2["A3"] = "Total sequences"
    ws2["B3"] = len(records)
    ws2["A4"] = "Total base pairs"
    ws2["B4"] = f"=SUM(Metadata!E2:E{len(records)+1})"

    # Genus counts
    from collections import Counter
    genus_counts = Counter(r.get("genus","Unknown") for r in records)
    ws2["A6"] = "Genus"
    ws2["B6"] = "Count"
    ws2["A6"].font = Font(bold=True, name="Arial")
    ws2["B6"].font = Font(bold=True, name="Arial")
    for i, (g, c) in enumerate(sorted(genus_counts.items()), 7):
        ws2[f"A{i}"] = g
        ws2[f"B{i}"] = c
        ws2[f"A{i}"].font = Font(name="Arial", size=10)
        ws2[f"B{i}"].font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12

    wb.save(out_path)


# ── FASTA parser ──────────────────────────────────────────────────────────────

def parse_fasta_records(fasta_text: str) -> list[dict]:
    """Parse multi-FASTA into list of {header, seq} dicts."""
    records = []
    current_header, current_seq = None, []
    for line in fasta_text.splitlines():
        if line.startswith(">"):
            if current_header is not None:
                records.append({"header": current_header,
                                 "seq": "".join(current_seq)})
            current_header = line[1:].strip()
            current_seq    = []
        else:
            current_seq.append(line.strip())
    if current_header:
        records.append({"header": current_header, "seq": "".join(current_seq)})
    return records


def extract_accession(header: str) -> str:
    """Extract accession from a FASTA header line."""
    parts = header.split()
    acc   = parts[0] if parts else header
    # Handle gb|ACC| format
    if "|" in acc:
        acc = acc.split("|")[1] if len(acc.split("|")) > 1 else acc
    return acc.split(".")[0]


# ── Metadata enrichment ───────────────────────────────────────────────────────

def enrich_with_taxonomy(summaries: list[dict]) -> list[dict]:
    """Extract family and genus from taxonomy field in esummary."""
    for s in summaries:
        title = s.get("title", "")
        org   = s.get("organism", "")
        # Try to infer genus from organism name (first word)
        if org and " " in org:
            s.setdefault("genus", org.split()[0])
        else:
            s.setdefault("genus", "")
        s.setdefault("family", "")
        # Try to extract host from title
        host_match = re.search(r"(?:isolated from|host[:\s]+)([^,;\.]+)", title, re.I)
        s.setdefault("host", host_match.group(1).strip() if host_match else "")
    return summaries


# ── Main download function ─────────────────────────────────────────────────────

def download_genomes(
    taxon:       str | None   = None,
    taxid:       str | None   = None,
    accessions:  list[str]    = None,
    limit:       int          = 100,
    complete:    bool         = True,
    output_dir:  str          = ".",
    prefix:      str          = "genome",
    verbose:     bool         = True,
    log_fn                    = None,
) -> dict:
    """
    Main download function. Returns summary dict.

    log_fn: optional callable(str) for streaming log messages.
    """
    def log(msg):
        if verbose: print(msg, flush=True)
        if log_fn:  log_fn(msg)

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Build search term ────────────────────────────────────────────────────
    if accessions:
        ids   = accessions
        term  = " OR ".join(f"{a}[Accession]" for a in accessions)
        log(f"Fetching {len(ids)} accessions...")
    else:
        # Resolve taxon name to txid if needed
        if taxon and not taxid:
            log(f"Resolving taxonomy: {taxon}...")
            taxid = taxon_to_txid(taxon)
            if taxid:
                log(f"  Taxonomy ID: {taxid}")
            else:
                log(f"  Could not resolve '{taxon}' — using as search term")
                taxid = None

        if taxid:
            term = f"txid{taxid}[Organism:exp]"
        elif taxon:
            term = f"{taxon}[Organism]"
        else:
            log("Error: provide --taxon, --taxid, or --accessions")
            return {"error": "no query"}

        if complete:
            term += " AND complete genome[Title]"

        log(f"Search term: {term}")

        # Search
        result = esearch(term, db="nucleotide", limit=limit)
        ids    = result.get("idlist", [])
        count  = result.get("count", "?")
        log(f"Found {count} sequences — downloading {len(ids)}")

        if not ids:
            log("No sequences found. Try a broader search or different taxon name.")
            return {"error": "no results", "count": 0}

    # ── Fetch sequences in batches ───────────────────────────────────────────
    BATCH  = 20
    all_fasta   = []
    all_summaries = []

    for i in range(0, len(ids), BATCH):
        batch = ids[i:i+BATCH]
        log(f"  Downloading sequences {i+1}-{min(i+BATCH, len(ids))} of {len(ids)}...")

        fasta_text = efetch_fasta(batch)
        records    = parse_fasta_records(fasta_text)
        all_fasta.extend(records)

        summaries = efetch_summary(batch)
        all_summaries.extend(summaries)

        time.sleep(0.4)  # NCBI rate limit

    log(f"Downloaded {len(all_fasta)} sequences")

    # ── Save individual FASTA files ──────────────────────────────────────────
    metadata_rows = []
    saved_files   = []

    for fi, rec in enumerate(all_fasta):
        acc      = extract_accession(rec["header"])
        filename = f"{prefix}_{acc}.fasta"
        filepath = out_dir / filename

        with open(filepath, "w") as f:
            f.write(f">{rec['header']}\n")
            # Wrap sequence at 70 chars
            seq = rec["seq"]
            for j in range(0, len(seq), 70):
                f.write(seq[j:j+70] + "\n")

        saved_files.append(str(filepath))

        # Build metadata row
        # Match with summary if available
        summ = {}
        for s in all_summaries:
            s_acc = s.get("accessionversion","").split(".")[0]
            if s_acc == acc or s.get("caption","") == acc:
                summ = s
                break

        metadata_rows.append({
            "accession":       acc,
            "organism":        summ.get("organism", ""),
            "family":          summ.get("family",   ""),
            "genus":           summ.get("genus",    ""),
            "length":          len(rec["seq"]),
            "gc":              gc_content(rec["seq"]),
            "title":           summ.get("title",    rec["header"][:80]),
            "collection_date": summ.get("subtype",  "").replace("collection_date/",""),
            "country":         summ.get("subname",  ""),
            "host":            "",
            "filename":        filename,
        })

        if (fi+1) % 10 == 0 or fi+1 == len(all_fasta):
            log(f"  Saved {fi+1}/{len(all_fasta)} files")

    # Enrich metadata
    metadata_rows = enrich_with_taxonomy(metadata_rows)

    # ── Save Excel metadata ──────────────────────────────────────────────────
    excel_path = str(out_dir / f"{prefix}_metadata.xlsx")
    log(f"Writing Excel metadata: {excel_path}")
    write_excel_metadata(metadata_rows, excel_path)
    saved_files.append(excel_path)

    # ── Summary ──────────────────────────────────────────────────────────────
    log(f"\nDone.")
    log(f"  Sequences:  {len(all_fasta)}")
    log(f"  Output dir: {out_dir}")
    log(f"  Metadata:   {excel_path}")

    return {
        "n_sequences": len(all_fasta),
        "files":       saved_files,
        "metadata":    metadata_rows,
        "excel":       excel_path,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

def merge_excel(input_files: list[str], output_path: str,
                verbose: bool = True) -> None:
    """
    Merge multiple kali_downloader metadata Excel files into one.
    Removes duplicate accessions, rebuilds the Summary sheet.
    """
    all_rows = []
    seen_accessions = set()

    for fpath in input_files:
        if not Path(fpath).exists():
            if verbose: print(f"  Skipping missing file: {fpath}")
            continue
        if verbose: print(f"  Reading: {fpath}")
        wb  = openpyxl.load_workbook(fpath, data_only=True)
        ws  = wb["Metadata"] if "Metadata" in wb.sheetnames else wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            rec = dict(zip(headers, row))
            acc = str(rec.get("Accession","")).strip()
            if acc and acc not in seen_accessions:
                seen_accessions.add(acc)
                # Normalise keys to match write_excel_metadata
                all_rows.append({
                    "accession":       rec.get("Accession",""),
                    "organism":        rec.get("Organism",""),
                    "family":          rec.get("Family",""),
                    "genus":           rec.get("Genus",""),
                    "length":          rec.get("Length (bp)",""),
                    "gc":              rec.get("GC %",""),
                    "title":           rec.get("Title / Description",""),
                    "collection_date": rec.get("Collection Date",""),
                    "country":         rec.get("Country",""),
                    "host":            rec.get("Host",""),
                    "filename":        rec.get("Filename",""),
                })

    if verbose:
        print(f"  Total unique sequences: {len(all_rows)}")

    write_excel_metadata(all_rows, output_path)
    if verbose:
        print(f"  Merged Excel saved: {output_path}")


def main():
    p = argparse.ArgumentParser(
        description="KALI Genome Downloader — browse and download from NCBI"
    )
    sub = p.add_subparsers(dest="command")

    # ── download command ──────────────────────────────────────
    dl = sub.add_parser("download", help="Download genomes from NCBI")
    dl.add_argument("--taxon",   nargs="+",
                    help="One or more taxon names (downloads each and merges into one Excel)")
    dl.add_argument("--taxid",   nargs="+",
                    help="One or more NCBI taxonomy IDs")
    dl.add_argument("--accessions", nargs="+", help="Specific accession numbers")
    dl.add_argument("--limit",   type=int, default=100,
                    help="Max sequences per taxon (default: 100)")
    dl.add_argument("--complete",action="store_true", default=True)
    dl.add_argument("--no-complete", action="store_true")
    dl.add_argument("--prefix",  default="genome")
    dl.add_argument("-o","--output", required=True)
    dl.add_argument("-v","--verbose", action="store_true")

    # ── merge command ─────────────────────────────────────────
    mg = sub.add_parser("merge", help="Merge multiple metadata Excel files into one")
    mg.add_argument("--inputs", nargs="+", required=True,
                    help="Excel files to merge")
    mg.add_argument("-o","--output", required=True,
                    help="Output merged Excel path")
    mg.add_argument("-v","--verbose", action="store_true")

    # ── backward compat: no subcommand = download ─────────────
    p.add_argument("--taxon",   nargs="+")
    p.add_argument("--taxid",   nargs="+")
    p.add_argument("--accessions", nargs="+")
    p.add_argument("--limit",   type=int, default=100)
    p.add_argument("--complete",action="store_true", default=True)
    p.add_argument("--no-complete", action="store_true")
    p.add_argument("--prefix",  default="genome")
    p.add_argument("-o","--output")
    p.add_argument("-v","--verbose", action="store_true")

    args = p.parse_args()

    # Merge command
    if args.command == "merge":
        merge_excel(args.inputs, args.output, args.verbose)
        return

    # Download command (with or without subcommand)
    complete   = not args.no_complete
    out_dir    = args.output
    prefix     = args.prefix
    limit      = args.limit
    verbose    = args.verbose
    accessions = args.accessions

    # Collect taxons/taxids
    taxons = args.taxon or []
    taxids = args.taxid or []

    if not taxons and not taxids and not accessions:
        p.print_help()
        sys.exit(1)

    # Single accession download
    if accessions:
        download_genomes(
            accessions = accessions,
            limit      = len(accessions),
            complete   = complete,
            output_dir = out_dir,
            prefix     = prefix,
            verbose    = verbose,
        )
        return

    # Multi-taxon: download each separately then merge Excel files
    all_excel_files = []
    all_result_files = []

    items = [(t, None) for t in taxons] + [(None, tid) for tid in taxids]

    for taxon, taxid in items:
        label  = taxon or f"txid{taxid}"
        safe   = label.replace(" ","_").replace("/","_")[:30]
        sub_dir = str(Path(out_dir) / safe)
        sub_pfx = prefix + "_" + safe

        print(f"\n── Downloading: {label} ──────────────────────────")
        result = download_genomes(
            taxon      = taxon,
            taxid      = taxid,
            limit      = limit,
            complete   = complete,
            output_dir = sub_dir,
            prefix     = sub_pfx,
            verbose    = verbose,
        )
        if "excel" in result:
            all_excel_files.append(result["excel"])
        if "files" in result:
            all_result_files.extend(result["files"])

    # Merge all Excel files if multiple taxons
    if len(all_excel_files) > 1:
        merged_path = str(Path(out_dir) / f"{prefix}_merged_metadata.xlsx")
        print(f"\n── Merging {len(all_excel_files)} Excel files ──────────────────────────")
        merge_excel(all_excel_files, merged_path, verbose=True)
        print(f"   Merged: {merged_path}")
    elif len(all_excel_files) == 1:
        print(f"\n   Excel: {all_excel_files[0]}")


if __name__ == "__main__":
    main()
